import openpyxl
import  jpype 
import  asposecells
jpype.startJVM()
from asposecells.api import SaveFormat, Workbook, FileFormatType

class excelLibro():

    def __init__(self, ruta):
        self.__ruta = ruta
        self.__libro = openpyxl.load_workbook(ruta, data_only=False)
        self.hoja = None
        self.__numMan = None
        self.__NameCond = None
        self.__idCond = None
        self.__fechaRep = None
        self.__placa = None
        self.__producto = None

    def extraer_datos(self, numMan):
        bool = False
        self.hoja = self.__libro['BDA1']
        celdas = self.hoja['A2':'AK'+str(self.hoja.max_row)]
        for fila in celdas:
            if fila[9].value == numMan:
                # VARIABLES DE SALIDA---------- #---0 nameCond--- 1 idCond --- 2 fechaRep --- 3 placa ---- 4 producto
                self.__numMan = fila[9].value
                self.__NameCond = fila[30].value
                self.__idCond = fila[29].value
                self.__fechaRep = fila[10].value
                self.__placa = fila[19].value
                self.__producto = fila[24].value
                bool = True
                break
        return bool

    def imprimir(self):
        print(f'nombre: {self.__NameCond} - idcond: {self.__idCond} - fecha: {self.__fechaRep} - placa: {self.__placa} - producto: {self.__producto} - numMan: {self.__numMan}')

    def guardar_info(self, rutaP):
        libroP = openpyxl.load_workbook(rutaP, data_only=False)
        
        self.hoja = libroP.active

        # llenar tabla

        self.hoja['B4'] = self.__numMan
        self.hoja['D4'] = self.__fechaRep
        self.hoja['F4'] = self.__placa
        self.hoja['B5'] = self.__NameCond
        self.hoja['D5'] = self.__idCond
        self.hoja['F5'] = self.__producto
        celdas = self.hoja['C48':'C50'] 
        for i in range(48,51):
            if self.__producto == 'CRUDO DE PALAMA' or self.__producto == 'CRUDO' or self.__producto == 'COMBUSTOLIO' or self.__producto == 'OLEINA' or self.__producto == 'ACEITES Y BASES VEGETALES' or self.__producto == 'COMBUSTOLEO' or self.__producto == 'ESTEARINA DE PALMA' :
                self.hoja['C'+str(i)] = 'N.A'
            else:
                self.hoja['C'+str(i)] = 'CONFORME'      
        celdas = self.hoja['F7':'F10'] 
        for i in range(7,11):
            if self.__producto == 'CRUDO DE PALAMA' or self.__producto == 'CRUDO' or self.__producto == 'COMBUSTOLIO' or self.__producto == 'OLEINA' or self.__producto == 'ACEITES Y BASES VEGETALES' or self.__producto == 'COMBUSTOLEO' or self.__producto == 'ESTEARINA DE PALMA' :
                self.hoja['F'+str(i)] = 'N.A'
            else:
                self.hoja['F'+str(i)] = 'CONFORME'
        celdas = self.hoja['F7':'F10'] 
        for i in range(12,49):
            if i != 32 :
                if self.__producto == 'CRUDO DE PALAMA' or self.__producto == 'CRUDO' or self.__producto == 'COMBUSTOLIO' or self.__producto == 'OLEINA' or self.__producto == 'ACEITES Y BASES VEGETALES' or self.__producto == 'COMBUSTOLEO' or self.__producto == 'ESTEARINA DE PALMA' :
                    self.hoja['F'+str(i)] = 'CONFORME'
                else:
                    self.hoja['F'+str(i)] = 'N.A'        
        # celdas=hoja['A2':'AK'+str(hoja.max_row)]
        rutaP='./archivo/print.xlsx'
        libroP.save(rutaP)
        return rutaP
    
    def crear_pdf(self,rutaXL):
        # Cargar archivo de Excel
        workbook  =  Workbook( rutaXL )
        rutaPDF = "./archivo/xlsx-to-pdf"+ str(self.__numMan) +".pdf"
        # Convertir Excel a PDF
        workbook.save( rutaPDF, SaveFormat.PDF )
        return rutaPDF





